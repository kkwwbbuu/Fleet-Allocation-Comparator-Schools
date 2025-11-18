import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st
from io import BytesIO

st.set_page_config(page_title="Schools Files Comparator", layout="centered")
st.title("Schools Files Comparator - Fleet Allocaiton")
st.write("Upload the Operations Report file and its corresponding Power Bi file to compare their contents.")

st.markdown("<h2 style='text-align:center; color:#27AE60;'>Upload SAP File</h2>", unsafe_allow_html=True)
st.write("Upload the SAP file as originally exported.")
uploaded_file1 = st.file_uploader("File 1", type=["xlsx", "xls", "xlsm", "xlsb"], key="file1")
st.markdown("<h2 style='text-align:center; color:#2E86C1;'>Upload Power BI File</h2>", unsafe_allow_html=True)
st.write("Upload the Power BI file as originally exported.")
uploaded_file2 = st.file_uploader("File 2", type=["xlsx", "xls", "xlsm", "xlsb"], key="file2")



if uploaded_file1 and uploaded_file2:
    try:
        file1 = pd.read_excel(uploaded_file1, sheet_name="Sheet1", dtype=str)
        file2 = pd.read_excel(uploaded_file2, dtype=str)
        print(file1.columns.tolist())
        print(file2.columns.tolist())

        if file1.columns[0] == "Fleet No":
            file1.rename(columns={
                "Fleet No": "Fleet Number",
                "Allocation Date": "Allocation date"
            }, inplace=True)

            # file1 = file1[
            #     (file1["Allocation master"].str.strip().str.lower() == "active") &
            #     (file1["PM Assign To"].str.strip().str.lower().isin(["psv", "metrolink", "metro express"]))
            # ]
        # Now select only the columns needed for comparison
        file1 = file1[["Fleet Number", "Allocation date", "Allocation Status", "Depot", "Assign To"]]
        file2 = file2[["Fleet Number", "Allocation date", "Allocation Status", "Depot", "Assigned to"]]

        file1_name = uploaded_file1.name
        file2_name = uploaded_file2.name
        merged = pd.merge(file1, file2, on="Fleet Number", how="outer", suffixes=('_file1', '_file2'))
        problems = []
        NumberErrors = 0
        StatusErrors = 0
        DateErrors = 0
        DepotErrors = 0
        AssignedErrors = 0
        error_IDs = set()

        for _, row in merged.iterrows():
            num = row["Fleet Number"]
            if pd.isna(num):
                continue
            if pd.isna(row["Allocation date_file1"]):
                problems.append(f"❌Fleet Number {num} is missing in file1_name")
                NumberErrors += 1
                error_IDs.add(num)
            elif pd.isna(row["Allocation date_file2"]):
                problems.append(f"❌Fleet Number {num} is missing in file2_name")
                NumberErrors += 1
                error_IDs.add(num)
            else:
                if row["Allocation date_file1"] != row["Allocation date_file2"]:
                    # problems.append(
                    #     f"⚠️Fleet Number {num} has different date "
                    #     f"(file1_name={row[' date_file1']}, file2_name={row[' date_file2']})"
                    # )
                    DateErrors += 1
                if row["Allocation Status_file1"] != row["Allocation Status_file2"]:
                    problems.append(
                        f"⚠️Fleet Number {num} has different Status "
                        f"({row['Allocation Status_file1']} vs {row['Allocation Status_file2']})"
                    )
                    StatusErrors += 1
                    error_IDs.add(num)
                if row["Depot_file1"] != row["Depot_file2"]:
                    problems.append(
                        f"⚠️Fleet Number {num} has different Depot "
                        f"({row['Depot_file1']} vs {row['Depot_file2']})"
                    )
                    DepotErrors += 1
                    error_IDs.add(num)
                if row["Assign To"] != row["Assigned to"]:
                    problems.append(
                        f"⚠️Fleet Number {num} has different Assigned To "
                        f"({row['Assign To']} vs {row['Assigned to']})"
                    )
                    AssignedErrors += 1
                    error_IDs.add(num)
                    
        valid_numbers = merged["Fleet Number"].dropna().unique()
        total_IDs = len(valid_numbers)
        total_error_IDs = len(error_IDs)
        if total_error_IDs > 0:
            percent_error = (total_error_IDs / total_IDs) * 100
            accuracy = ((total_IDs - total_error_IDs) / total_IDs) * 100
        else:
            percent_error = 0
            accuracy = 100
        
        summary_table = [
            ["Accuracy", f"{accuracy:.2f}%", ""],
            ["Fleet Numbers with mismatches", f"{percent_error:.2f}%", ""],
            ["Fleet Number missing case(s)", NumberErrors, ""],
            ["Allocation Status mismatch case(s)", StatusErrors, ""],
            ["Allocation Depot mismatch case(s)", DepotErrors, ""],
            ["Assigned To mismatch case(s)", AssignedErrors, ""]

        ]

        # #For displaying the summary
        # st.subheader("📈 Summary")
        # st.write(f"**Accuracy:** {accuracy:.2f}%")
        # st.write(f"**Numbers with mismatches:** {percent_error:.2f}%")
        # st.write(f"**Number missing case(s):** {NumberErrors}")
        # st.write(f"**Status mismatch case(s):** {StatusErrors}")

        # if problems:
        #     st.warning("⚠️ Issues found:")
        #     for p in problems:
        #         st.text(p)
        # else:
        #     st.success("✅ No problems found!")

        analytics_data = []
        for _, row in merged.iterrows():
            num = row["Fleet Number"]
            if pd.isna(num):
                continue
            if pd.isna(row["Allocation date_file1"]):
                analytics_data.append({
                    "Fleet Number": num,
                    "Depot in SAP" : row["Depot_file1"],
                    "Assigned To in SAP" : row["Assign To"],
                    "Type": "Missing Number",
                    f"{file1_name}": "Missing",
                    f"{file2_name}": ""
                })
            elif pd.isna(row["Allocation date_file2"]):
                analytics_data.append({
                    "Fleet Number": num,
                    "Depot in SAP" : row["Depot_file1"],
                    "Assigned To in SAP" : row["Assign To"],
                    "Type": "Missing Number",
                    f"{file1_name}": "",
                    f"{file2_name}": "Missing"
                })
            elif row["Allocation Status_file1"] != row["Allocation Status_file2"]:
                analytics_data.append({
                    "Fleet Number": num,
                    "Depot in SAP" : row["Depot_file1"],
                    "Assigned To in SAP" : row["Assign To"],
                    "Type": "Status Mismatch",
                    f"{file1_name}": row["Allocation Status_file1"],
                    f"{file2_name}": row["Allocation Status_file2"]
                })
            elif row["Depot_file1"] != row["Depot_file2"]:
                analytics_data.append({
                    "Fleet Number": num,
                    "Depot in SAP" : row["Depot_file1"],
                    "Assigned To in SAP" : row["Assign To"],
                    "Type": "Depot Mismatch",
                    f"{file1_name}": row["Depot_file1"],
                    f"{file2_name}": row["Depot_file2"]
                })
            elif row["Assign To"] != row["Assigned to"]:
                analytics_data.append({
                    "Fleet Number": num,
                    "Depot in SAP" : row["Depot_file1"],
                    "Assigned To in SAP" : row["Assign To"],
                    "Type": "Assigned To Mismatch",
                    f"{file1_name}": row["Assign To"],
                    f"{file2_name}": row["Assigned to"]
                })

        df_comparison = pd.DataFrame(analytics_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = "Comparison Result"
            
            # ✅ Write df_comparison first to create the sheet
            df_comparison.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_table) + 3)
            
            worksheet = writer.sheets[sheet_name]
            
            # ✅ Now safely modify the sheet
            worksheet.merge_cells('A1:D1')
            cell = worksheet['A1']
            cell.value = "Comparison Result"
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center")

            # Write summary table
            for i, row in enumerate(summary_table, start=3):
                for j, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=i, column=j, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Adjust column widths
            for i, col in enumerate(worksheet.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = max_length + 4

            # Center-align analytics rows
            startrow = len(summary_table) + 3
            for row in worksheet.iter_rows(
                min_row=startrow,
                max_row=startrow + len(df_comparison) + 1,
                min_col=1,
                max_col=len(df_comparison.columns)
            ):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")


        st.markdown("<h3 style='color:#FF0000; font-size:28px;'>Enter file name for download</h3>", unsafe_allow_html=True)
        filename = st.text_input("File name for download", value="Schools_fleet_allocation_comparison_result.xlsx")

        st.download_button(
            label="💾 Download Comparison Excel",
            data=output.getvalue(),
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")